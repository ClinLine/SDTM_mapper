import jsonata
import json
import openpyxl
import pandas as pd

# Define the source json file you like to use
JsonInput = "TestJson/EliLilly_NCT03421379_Diabetes.json"

wb = openpyxl.load_workbook("Maps/sdtm_mapping_paths.xlsx")

# Access the 'TS Parameters' sheet
ts0_sheet = wb['TS']
for i in range(2, ts0_sheet.max_row + 1):
    j=i-1
    #swap the rows and columns in the TS sheet
    varName=ts0_sheet.cell(row=i, column=1).value
    ts0_sheet.cell(row=1, column=j).value = varName
    if varName == "STUDYID":
        StudyIdCodeSnip = ts0_sheet.cell(row=i, column=7).value
    if varName == "DOMAIN":
        DomainResult =  ts0_sheet.cell(row=i, column=8).value


ts_sheet = wb['TS Parameters']

# Print the value in the first and seventh column of each row in the 'TS Parameters' sheet
with open(JsonInput, 'r') as file:
    data=json.load (file)
    # Get StudyId first and start with a row id for the TS sheet
    expr = jsonata.Jsonata (StudyIdCodeSnip)
    studyId= expr.evaluate(data)
    x=1
    # Then continue with the mappings in the TS Parameters sheet
    ts_sheet.cell(row=1, column=7).value = "Mapping Results"
    for i in range(2, ts_sheet.max_row + 1):
        # Get all the mapping information from the TS Parameters sheet
        MapName = ts_sheet.cell(row=i, column=1).value
        MapCode = ts_sheet.cell(row=i, column=2).value
        codeSnip = ts_sheet.cell(row=i, column=7).value        
        nfValue = ts_sheet.cell(row=i, column=8).value
        codeSnipCd = ts_sheet.cell(row=i, column=9).value   
        codeSnipCdRef = ts_sheet.cell(row=i, column=10).value   
        codeSnipCdVer = ts_sheet.cell(row=i, column=11).value   

        resultCd=" "
        resultCdRef=" "
        resultCdVer=" "
        if codeSnip is None:
            result=" "
        else:
            # print(codeSnip)
            try:
                expr = jsonata.Jsonata(codeSnip)
                result = expr.evaluate(data)  
            except:
                result = "Error in expression for "+ MapName + ": " + codeSnip
        if codeSnipCd is not None:
            try:
                exprCd = jsonata.Jsonata(codeSnipCd)
                resultCd = exprCd.evaluate(data)  
            except:
                resultCd = "Error in expression for "+ MapName + ": " + codeSnipCd
        if codeSnipCdRef is not None:
            try:
                exprCdRef = jsonata.Jsonata(codeSnipCdRef)
                resultCdRef = exprCdRef.evaluate(data)  
            except:
                resultCdRef = "Error in expression for "+ MapName + ": " + codeSnipCdRef
        if codeSnipCdVer is not None:
            try:
                exprCdVer = jsonata.Jsonata(codeSnipCdVer)
                resultCdVer = exprCdVer.evaluate(data)  
            except:
                resultCdVer = "Error in expression for "+ MapName + ": " + codeSnipCdVer
        
       
        if result is None: result = " "
        if nfValue is None: nfValue = " "
        if resultCd is None: resultCd = " "
        if resultCdRef is None: resultCdRef = " "
        if resultCdVer is None: resultCdVer = " "
        # print(result)
        result= str(result)
        resultCd=str(resultCd)
        resultCdRef=str(resultCdRef)
        resultCdVer=str(resultCdVer)
        # replace the apostrophes with spaces
        try:
            result2 = result.replace("’", " ")
        except:
            result2 = None
        if result2 is None: result2= " "
        if result2 == "": result2= " "
        if result2 == "{}": result2 = " "
        if result2[0] == "[": 
            result2 = result2[1:-1]
            result2 = result2.replace("}, {", " , ")
        if result2 != " ":
            if result2[0] == "{":  # check if the result is a list
                print (result2)
                n = 0 #letter it is looking at
                result3 = []
                while result2[n] != "}":
                    if result2[n] == "{" or ",":
                        n += 1
                        m = n
                        while m+1 < len(result2) and result2[m+1] not in ("}", ","):
                            m += 1
                        result3.append(result2[n:m])
                        n = m + 1
                    else: 
                        n += 1
                print(result3)   
                for j in range(len(result3)):
        # filling TS Parameters sheet
        ts_sheet.cell(row=i, column=7).value = result2
        ts_sheet.cell(row=i, column=8).value = " "   
        if result2 == " ": ts_sheet.cell(row=i, column=8).value = nfValue	
        ts_sheet.cell(row=i, column=9).value = resultCd  
        ts_sheet.cell(row=i, column=10).value = resultCdRef
        ts_sheet.cell(row=i, column=11).value = resultCdVer
        # filling TS sheet
        if result2 != " " or nfValue != " ":
            x=x+1
            ts0_sheet.cell(row=x, column=1).value = " "
            ts0_sheet.cell(row=x, column=1).value = studyId
            ts0_sheet.cell(row=x, column=2).value = DomainResult   
            ts0_sheet.cell(row=x, column=3).value = " "
            ts0_sheet.cell(row=x, column=4).value = " "
            ts0_sheet.cell(row=x, column=5).value = MapCode    
            ts0_sheet.cell(row=x, column=6).value = MapName                
            ts0_sheet.cell(row=x, column=7).value = result2   
            ts0_sheet.cell(row=x, column=8).value = " "   
            if result2 == " ": ts0_sheet.cell(row=x, column=8).value = nfValue	
            ts0_sheet.cell(row=x, column=9).value = resultCd                   
            ts0_sheet.cell(row=x, column=10).value = resultCdRef  
            ts0_sheet.cell(row=x, column=11).value = resultCdVer
        
    file.close
wb.save("Output/sdtm_mapping_results.xlsx")